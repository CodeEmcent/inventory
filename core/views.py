from rest_framework.viewsets import ModelViewSet
from rest_framework import serializers, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError, PermissionDenied
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import NotFound
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination
import csv
from io import StringIO
from openpyxl import load_workbook
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Avg, Count
from datetime import date
from .models import Office, ItemRegister, InventoryItem
from .serializers import (
    OfficeSerializer,
    ItemRegisterSerializer,
    InventoryItemSerializer,
)
from accounts.permissions import (
    IsAdminOrStaffOrReadOnly,
    IsAdminOrSuperAdmin,
    IsOwnerOrAdminOrStaff,
    IsAssignedStaff,
    IsAssignedStaffOrReadOnly,
)

class InventoryPagination(PageNumberPagination):
    page_size = 15  # Number of items per page
    page_size_query_param = "page_size"  # Allow client to specify page size
    max_page_size = 100  # Limit the maximum size to prevent abuse

# --- Office ViewSet ---
class OfficeViewSet(ModelViewSet):
    queryset = Office.objects.all()
    serializer_class = OfficeSerializer
    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin | IsAssignedStaffOrReadOnly]
    def get_queryset(self):
        """
        Restrict queryset to offices assigned to the user for staff.
        Admins and superadmins can view all offices.
        """
        user = self.request.user
        if user.role == "staff":
            # Return only offices assigned to the staff user
            return Office.objects.filter(assigned_users=user)
        # Return all offices for admins and superadmins
        return super().get_queryset()

    def perform_create(self, serializer):
        """
        Ensure staff users can only create objects for their assigned offices.
        """
        user = self.request.user
        if user.role == 'staff' and serializer.validated_data.get("office") not in user.assigned_offices.all():
            raise PermissionDenied("You do not have permission to create inventory for this office.")
        serializer.save()

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        response.data["message"] = "Office created successfully."
        return response

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        response.data["message"] = "Office updated successfully."
        return response

    def destroy(self, request, *args, **kwargs):
        office = self.get_object()
        super().destroy(request, *args, **kwargs)
        return Response(
            {"message": f"Office '{office.name}' deleted successfully."}, status=200
        )

class ItemRegisterViewSet(ModelViewSet):
    """
    Handles CRUD operations for the Item Register.
    """

    queryset = ItemRegister.objects.all()
    serializer_class = ItemRegisterSerializer
    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin | IsAssignedStaffOrReadOnly]
    lookup_field = "item_id"  # Use 'item_id' for lookups instead of the default 'id'

    def get_object(self):
        """
        Override the get_object method to look up by item_id instead of id.
        """
        item_id = self.kwargs.get("item_id")
        try:
            return ItemRegister.objects.get(item_id=item_id)
        except ItemRegister.DoesNotExist:
            raise NotFound(detail="Item not found")

    def list(self, request, *args, **kwargs):
        """
        Allow all users to view the item register, but restrict editing to admins.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)  # Use DRF pagination if needed
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        data = [
            {
                "id": item.id,
                "item_id": item.item_id,
                "name": item.name,
                "description": item.description or "N/A",
            }
            for item in queryset
        ]
        return Response({"item_register": data}, status=200)

    def create(self, request, *args, **kwargs):
        """
        Create a new item in the Register (Restricted to Admins and Superadmins).
        """
        self.check_permissions(request)  # Ensure only admin/superadmin can create items
        response = super().create(request, *args, **kwargs)
        response.data["message"] = "Item registered successfully."
        return response

    def update(self, request, *args, **kwargs):
        """
        Update an item in the Register (Restricted to Admins and Superadmins).
        """
        self.check_permissions(request)  # Ensure only admin/superadmin can update items
        response = super().update(request, *args, **kwargs)
        response.data["message"] = "Item updated successfully."
        return response

    def destroy(self, request, *args, **kwargs):
        """
        Delete an item from the Register (Restricted to Admins and Superadmins).
        """
        self.check_permissions(request)  # Ensure only admin/superadmin can delete items
        item = self.get_object()
        super().destroy(request, *args, **kwargs)
        return Response(
            {"message": f"Item '{item.name}' deleted successfully."}, status=200
        )

class RegisterTemplateView(APIView):
    """
    Endpoint to download an Excel template for item Register import.
    """

    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin]

    def get(self, request):
        # Create a workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Item Register Template"

        # Add organization name and header
        organization_name = (
            request.user.organization.name
            if request.user.organization
            else "Unknown Organization"
        )
        current_year = date.today().year  # Get the current year
        header_font = Font(bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Organization Name
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        org_cell = sheet.cell(row=1, column=1)
        org_cell.value = organization_name
        org_cell.font = Font(bold=True, size=14)
        org_cell.alignment = centered_alignment

        # Subheading: Year
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        year_cell = sheet.cell(row=2, column=1)
        year_cell.value = f"Item Register Template - Year {current_year}"
        year_cell.font = Font(bold=True, italic=True, size=12)
        year_cell.alignment = centered_alignment

        # Column headers
        headers = ["S/N", "Name", "Description"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename=item_Register_template.xlsx"
        )
        workbook.save(response)
        return response

class RegisterImportView(APIView):
    """
    Endpoint to upload an Excel template and register items in the Register.
    """

    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin]

    def post(self, request):
        # Get the uploaded file
        file_obj = request.FILES.get("file", None)
        if not file_obj:
            return Response({"error": "No file uploaded."}, status=400)

        try:
            # Load the workbook and sheet
            workbook = load_workbook(file_obj)
            sheet = workbook.active
        except Exception as e:
            return Response({"error": "Invalid Excel file format."}, status=400)

        # Validate headers
        expected_headers = ["S/N", "Name", "Description"]
        actual_headers = [cell.value for cell in sheet[3]]
        if actual_headers[: len(expected_headers)] != expected_headers:
            return Response(
                {"error": "Invalid template format. Please use the correct template."},
                status=400,
            )

        # Process rows
        created_items = []
        updated_items = []
        errors = []
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            serial_number, name, description = row

            # Validate row data
            if not name:
                errors.append(f"Row {i}: Name is required.")
                continue

            # Create or update item
            try:
                item, created = ItemRegister.objects.get_or_create(
                    name=name.strip(),
                    defaults={
                        "description": description.strip() if description else None
                    },
                )
                if created:
                    created_items.append(item.name)
                else:
                    # Update the description if provided
                    if description:
                        item.description = description.strip()
                        item.save()
                    updated_items.append(item.name)
            except Exception as e:
                errors.append(
                    f"Row {i}: Error processing item '{name}'. Details: {str(e)}"
                )

        # Prepare response
        response_data = {
            "message": "Import completed.",
            "created_items": created_items,
            "updated_items": updated_items,
        }
        if errors:
            response_data["errors"] = errors

        return Response(response_data, status=201 if not errors else 400)

class RegisterDownloadView(APIView):
    """
    Endpoint to download the item Register as an Excel file.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Create a workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Item Register"

        # Add organization name and header
        organization_name = (
            request.user.organization.name
            if request.user.organization
            else "Unknown Organization"
        )
        current_year = date.today().year  # Get the current year
        header_font = Font(bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Organization Name
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        org_cell = sheet.cell(row=1, column=1)
        org_cell.value = organization_name
        org_cell.font = header_font
        org_cell.alignment = centered_alignment

        # Subheading: Year
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        year_cell = sheet.cell(row=2, column=1)
        year_cell.value = f"Item Register - Year {current_year}"
        year_cell.font = Font(bold=True, italic=True, size=12)
        year_cell.alignment = centered_alignment

        # Column headers
        headers = ["S/N", "Item ID", "Name", "Description"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Fetch item Register data and populate the sheet
        item_Register = ItemRegister.objects.all()
        for idx, item in enumerate(item_Register, start=1):
            sheet.append(
                [
                    idx,  # S/N
                    item.item_id,  # item ID
                    item.name,  # Name
                    item.description or "N/A",  # Description
                ]
            )

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=item_Register.xlsx"
        workbook.save(response)
        return response

# --- Inventory ViewSet ---
class InventoryViewSet(ModelViewSet):
    queryset = InventoryItem.objects.all()
    serializer_class = InventoryItemSerializer
    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin | IsAssignedStaffOrReadOnly]
    pagination_class = InventoryPagination

    def get_queryset(self):
        """
        Restrict queryset based on user's role and assigned offices.
        """
        user = self.request.user
        office_id = self.request.query_params.get("office_id")

        if user.role == "staff":
            # Get all inventory for staff user's assigned offices if no `office_id` is provided
            assigned_offices = user.assigned_offices.all()
            if not assigned_offices.exists():
                raise PermissionDenied("You are not assigned to any office.")
            
            if not assigned_offices.exists():
                raise PermissionDenied("You are not assigned to any office.")

            if office_id:
                # Ensure the specified office is in the user's assigned offices
                office = get_object_or_404(Office, id=office_id)
                if office not in assigned_offices:
                    raise PermissionDenied(
                        "You do not have permission to view this office's inventory."
                    )
                return InventoryItem.objects.filter(office=office)

            # Default to all assigned offices
            return InventoryItem.objects.filter(office__in=assigned_offices)

        # Admins and superadmins can view all inventory items
        return InventoryItem.objects.all()

    def perform_create(self, serializer):
        """
        Handle creation of inventory items, ensuring staff users are associating with their assigned office.
        """
        user = self.request.user

        # Validate and fetch office
        office_id = self.request.data.get("office_id")
        if not office_id:
            raise ValidationError("Office ID is required.")

        office = get_object_or_404(Office, id=office_id)

        if user.role == "staff" and office not in user.assigned_offices.all():
            raise ValidationError("You are not assigned to this office.")

        # Validate and fetch item
        item_id = self.request.data.get("item_id")
        if not item_id:
            raise ValidationError("Item ID is required.")

        item = get_object_or_404(ItemRegister, item_id=item_id)

        # Save the inventory item with the correct argument name (item_id instead of item)
        serializer.save(user=user, office=office, item_id=item)  # Corrected to item_id instead of item

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # Restrict staff users to assigned offices
        user = self.request.user
        if user.role == "staff" and instance.office not in user.assigned_offices.all():
            raise PermissionDenied("You do not have permission to delete this item.")

        # Perform deletion
        self.perform_destroy(instance)

        return Response(
            {"message": f"Item '{instance.item_id}' successfully deleted."},
            status=200,
        )


# --- Template View ---
class TemplateView(APIView):
    permission_classes = [IsAuthenticated, IsAssignedStaff]

    def get(self, request, office_id):
        office = get_object_or_404(Office, id=office_id)

        # Authorization check for the office
        if (
            request.user.role not in ("admin", "super_admin")
            and office not in request.user.assigned_offices.all()
        ):
            return Response(
                {"error": "You do not have permission to download this template."},
                status=403,
            )

        # Fetch all items from the ItemRegister
        items = ItemRegister.objects.all()

        # Create workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"Template for {office.name}"

        # Header Font and Alignment
        header_font = Font(bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Organization Name Header
        organization_name = (
            request.user.organization.name
            if request.user.organization
            else "Unknown Organization"
        )
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        sheet.cell(row=1, column=1).value = organization_name
        sheet.cell(row=1, column=1).font = header_font
        sheet.cell(row=1, column=1).alignment = centered_alignment

        # Office Description Header
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        office_description = f"Office: {office.name} | Description: {office.department or 'No Description'}"
        sheet.cell(row=2, column=1).value = office_description
        sheet.cell(row=2, column=1).font = header_font
        sheet.cell(row=2, column=1).alignment = centered_alignment

        # Column Headers
        headers = [
            "S/N",
            "Item ID",
            "Items",
            "Qty",
            "Description (Optional)",
            "Remarks",
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Populate rows with items from ItemRegister, including the description
        for idx, item in enumerate(items, start=1):
            sheet.append([idx, item.item_id, item.name, "", item.description, ""])

        # Footer for Staff Name
        staff_name = request.user.username
        sheet.append([])  # Leave a blank row
        sheet.append(
            [
                f"Signature: {staff_name}",
                "Ensure item ID matches the uploaded template.",
            ]
        )
        sheet.merge_cells(
            start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=6
        )
        sheet.cell(row=sheet.max_row, column=1).alignment = centered_alignment

        # Adjust column widths
        column_widths = [10, 20, 30, 10, 40, 30]
        for col_num, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + col_num)].width = width

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename={office.name}_template.xlsx"
        )
        workbook.save(response)
        return response

# --- Import Inventory View ---
class ImportInventoryView(APIView):
    """
    Endpoint to import inventory data from an Excel file.
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        file_obj = request.FILES.get("file", None)
        office_id = request.query_params.get("office_id")

        office_id = request.query_params.get("office_id")

        if not file_obj:
            return Response({"error": "No file uploaded."}, status=400)
        if not office_id:
            return Response({"error": "Office ID is required."}, status=400)

        office = get_object_or_404(Office, id=office_id)
        if office not in request.user.assigned_offices.all():
            return Response(
                {"error": "You do not have permission to manage this office."},
                status=403,
            )

        workbook = load_workbook(file_obj)
        sheet = workbook.active
        if not office_id:
            return Response({"error": "Office ID is required."}, status=400)

        office = get_object_or_404(Office, id=office_id)
        if office not in request.user.assigned_offices.all():
            return Response(
                {"error": "You do not have permission to manage this office."},
                status=403,
            )

        workbook = load_workbook(file_obj)
        sheet = workbook.active

        # Validate headers
        expected_headers = [
            "S/N",
            "Item ID",
            "Items",
            "Qty",
            "Description (Optional)",
            "Remarks",
        ]
        actual_headers = [
            str(cell.value).strip() for cell in sheet[3]
        ]  # Normalize headers
        if actual_headers != expected_headers:
            return Response(
                {
                    "error": "Invalid template format. Please use the correct template.",
                    "expected_headers": expected_headers,
                    "actual_headers": actual_headers,  # Include actual headers for debugging
                },
                status=400,
            )

        # Process rows, starting from row 4
        imported_items = []
        updated_items = []
        for i, row in enumerate(
            sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True)
        ):
            serial_number, item_id, item_name, quantity, description, remarks = row

            # Skip rows where item_id is empty (such as empty or signature rows)
            if not item_id:
                continue

            # Validate item ID
            item = ItemRegister.objects.filter(item_id=item_id).first()
            if not item:
                return Response(
                    {"error": f"Invalid item ID '{item_id}' in row {i + 4}."},
                    status=400,
                )

            # Ensure item ID matches item name
            if item.name != item_name:
                return Response(
                    {
                        "error": f"Item name '{item_name}' does not match item ID '{item_id}' in row {i + 4}."
                    },
                    status=400,
                )

            # Validate quantity
            if not isinstance(quantity, int) or quantity < 1:
                return Response(
                    {
                        "error": f"Invalid quantity '{quantity}' in row {i + 4}. Must be a positive integer."
                    },
                    status=400,
                )

            # Process inventory item
            existing_item = InventoryItem.objects.filter(
                user=request.user, office=office, item_id=item
            ).first()
            if existing_item:
                existing_item.quantity = quantity
                existing_item.remarks = remarks or ""
                existing_item.save()
                updated_items.append(existing_item.item_id.name)
            else:
                imported_items.append(
                    InventoryItem(
                        user=request.user,
                        office=office,
                        item_id=item,
                        quantity=quantity,
                        remarks=remarks or "",
                        description=description or "",  # Add the description field
                    )
                )

        # Bulk create new inventory items
        InventoryItem.objects.bulk_create(imported_items)

        return Response(
            {
                "message": "Inventory imported successfully.",
                "new_items": len(imported_items),
                "updated_items": updated_items,
            },
            status=201,
        )

# --- Export Inventory View ---
class ExportInventoryView(APIView):
    """
    View to export all inventory items to an Excel file with enhanced structure, including item_id.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request):
        office_id = request.query_params.get("office_id")

        # Debug: Check the office_id value being passed
        print("Office ID:", office_id)

        # For staff users, apply office-related restrictions
        if request.user.role == "staff":
            if not office_id:
                return Response(
                    {"error": "Office ID is required to export inventory."}, status=400
                )

            office = Office.objects.filter(
                id=office_id, assigned_users=request.user
            ).first()

            # Debug: Check if the office is found for staff
            print("Staff Office:", office)

            if not office:
                return Response(
                    {
                        "error": "You do not have permission to export inventory for this office."
                    },
                    status=403,
                )

            inventory_items = InventoryItem.objects.filter(
                user=request.user, office=office
            )

        else:
            # Admins and Superadmins can export inventory for all offices
            if office_id:
                # Check if the office exists
                office = Office.objects.filter(id=office_id).first()

                # Debug: Check if the office exists for Admin/Superadmin
                print("Admin/Superadmin Office:", office)

                if not office:
                    return Response(
                        {"error": "The specified office does not exist."}, status=404
                    )

                # Admins/Superadmins can export inventory for that specific office
                inventory_items = InventoryItem.objects.filter(office=office)
            else:
                # If no office_id is provided, admins and superadmins can access all inventory items
                inventory_items = InventoryItem.objects.all()
                office = None  # Admins/Superadmins can export for all offices

        # Create the workbook and worksheet
        # Create the workbook and worksheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Inventory Items"

        # Header Font and Alignment
        header_font = Font(bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Add organization name as the first row
        # Header Font and Alignment
        header_font = Font(bold=True, size=14)
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Add organization name as the first row
        organization_name = (
            request.user.organization.name
            if request.user.organization
            else "Unknown Organization"
        )
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        sheet.cell(row=1, column=1).value = organization_name
        sheet.cell(row=1, column=1).font = header_font
        sheet.cell(row=1, column=1).alignment = centered_alignment

        # Add office name and department as the second row
        if office:
            office_details = f"Office: {office.name} | Department: {office.department or 'No Department'}"
        else:
            office_details = "All Offices (Admin/Superadmin Export)"
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        sheet.cell(row=2, column=1).value = office_details
        sheet.cell(row=2, column=1).font = header_font
        sheet.cell(row=2, column=1).alignment = centered_alignment

        # Add column headers, including item ID
        headers = [
            "S/N",
            "Item ID",
            "Item Name",
            "Quantity",
            "Description",
            "Remarks",
            "Created At",
            "Updated At",
        ]
        sheet.append(headers)

        # Style the column headers
        # Style the column headers
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_num)
            cell = sheet.cell(row=3, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows, including descriptions
        for idx, item in enumerate(inventory_items, start=1):
            sheet.append(
                [
                    idx,
                    item.item_id.item_id,  # item ID from the linked ItemRegister
                    item.item_id.name,  # Name from the linked ItemRegister
                    item.quantity,
                    item.description or "N/A",  # Include description in the row
                    item.remarks or "N/A",
                    item.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    item.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )

        # Add staff name as the footer
        staff_name = request.user.username
        sheet.append([])  # Leave a blank row
        sheet.append([f"Exported by: {staff_name}"])
        sheet.merge_cells(
            start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=8
        )
        sheet.cell(row=sheet.max_row, column=1).alignment = centered_alignment

        # Adjust column widths for better readability
        column_widths = [10, 15, 30, 10, 30, 30, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = (
                width  # Convert 1 to 'A', 2 to 'B', etc.
            )

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=inventory_items.xlsx"
        workbook.save(response)
        return response

class BroadsheetView(APIView):
    """
    API to generate a detailed broadsheet report with proper department and office mapping, including item_id.
    """

    permission_classes = [IsAuthenticated, IsAdminOrSuperAdmin]

    def get(self, request, *args, **kwargs):
        year = request.GET.get("year", None)
        if not year:
            return Response({"error": "Year parameter is required."}, status=400)

        # Aggregate inventory data and department-office mapping
        inventory_data = self.aggregate_inventory_data(year)
        department_offices = self.get_department_offices(year)

        # Generate Excel file
        return self.generate_excel(inventory_data, department_offices, year)

    def aggregate_inventory_data(self, year):
        """
        Fetch and group inventory data by unique items across offices, including item_id and descriptions.
        """
        return (
            InventoryItem.objects.filter(year=year)
            .values(
                "item_id__item_id",  # item ID from ItemRegister
                "item_id__name",  # Item name from ItemRegister
                "office__name",  # Office name
                "office__department",  # Department
                "item_id__description",  # Item description
                "item_id__unit_cost",  # Unit cost from ItemRegister
            )
            .annotate(total_quantity=Sum("quantity"))
            .distinct()
            .order_by("item_id__name")  # Order items alphabetically by name
        )

    def get_department_offices(self, year):
        """
        Fetch all unique departments and their associated offices without duplicates.
        """
        departments = (
            InventoryItem.objects.filter(year=year)
            .values_list("office__department", flat=True)
            .distinct()
        )
        department_offices = {}
        for dept in departments:
            offices = (
                InventoryItem.objects.filter(office__department=dept, year=year)
                .values_list("office__name", flat=True)
                .distinct()
            )
            department_offices[dept] = sorted(set(offices))  # Ensure unique and sorted
        return department_offices

    def generate_excel(self, inventory_data, department_offices, year):
        """
        Generate and return the broadsheet as an Excel file.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Broadsheet"

        # Organization and subheading
        # Fetch the organization name dynamically from user profile or directly from user
        organization_name = (
            getattr(self.request.user.profile, "organization_name", None)
            if hasattr(self.request.user, "profile")
            else None
        )
        if not organization_name:
            organization = (
                self.request.user.organization
                if hasattr(self.request.user, "organization")
                else None
            )
            organization_name = (
                organization.name if organization else "Unknown Organization"
            )
        num_office_columns = sum(
            len(offices) for offices in department_offices.values()
        )
        num_columns = 7 + num_office_columns  # Updated for two new columns

        # Header: Organization Name
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=num_columns
        )
        sheet.cell(row=1, column=1, value=organization_name.upper()).font = Font(
            size=21, bold=True
        )
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Subheading: Year
        sheet.merge_cells(
            start_row=2, start_column=1, end_row=2, end_column=num_columns
        )
        sheet.cell(
            row=2, column=1, value=f"Inventory Data for the Year {year}"
        ).font = Font(size=17, italic=True)
        sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        # Headers
        sheet.append(
            ["S/N", "ITEM ID", "ITEM NAME", "DESCRIPTION", "TOTAL"]
            + [
                f"{office}"
                for offices in department_offices.values()
                for office in offices
            ]
            + ["UNIT", "TOTAL VALUE"]
        )
        header_row = sheet.max_row

        # Apply bold font to all cells in the header row
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)

        # Merge cells for the label rows
        for col in range(1, 5):
            sheet.merge_cells(
                start_row=header_row,
                start_column=col,
                end_row=header_row + 1,
                end_column=col,
            )
            sheet.cell(row=header_row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Add department and office headers
        col_index = 5
        for department, offices in department_offices.items():
            start_col = col_index
            for office in offices:
                sheet.cell(row=header_row + 1, column=col_index, value=office)
                sheet.cell(row=header_row + 1, column=col_index).alignment = Alignment(
                    textRotation=90, horizontal="center"
                )
                col_index += 1
            sheet.merge_cells(
                start_row=header_row,
                start_column=start_col,
                end_row=header_row,
                end_column=col_index - 1,
            )
            sheet.cell(row=header_row, column=start_col).value = department
            sheet.cell(row=header_row, column=start_col).alignment = Alignment(
                horizontal="center", vertical="center"
            )

        # Format TOTAL, UNIT, TOTAL VALUE labels
        total_cell = sheet.cell(row=header_row, column=col_index, value="TOTAL")
        total_cell.font = Font(name="Times New Roman", size=12, bold=True)
        total_cell.alignment = Alignment(textRotation=90, horizontal="center")
        sheet.merge_cells(
            start_row=header_row,
            start_column=col_index,
            end_row=header_row + 1,
            end_column=col_index,
        )

        unit_cell = sheet.cell(row=header_row, column=col_index + 1, value="UNIT")
        unit_cell.font = Font(name="Times New Roman", size=12, bold=True)
        unit_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(
            start_row=header_row,
            start_column=col_index + 1,
            end_row=header_row + 1,
            end_column=col_index + 1,
        )

        total_value_cell = sheet.cell(
            row=header_row, column=col_index + 2, value="VALUE"
        )
        total_value_cell.font = Font(name="Times New Roman", size=12, bold=True)
        total_value_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(
            start_row=header_row,
            start_column=col_index + 2,
            end_row=header_row + 1,
            end_column=col_index + 2,
        )

        # Write data rows
        unique_items = inventory_data.values(
            "item_id__item_id",
            "item_id__name",
            "item_id__description",
            "item_id__unit_cost",
        ).distinct()
        row_index = sheet.max_row + 1
        serial_number = 1
        for item in unique_items:
            item_id = item["item_id__item_id"]
            item_name = item["item_id__name"]
            description = item["item_id__description"]
            unit_cost = item["item_id__unit_cost"] or 0

            # Write S/N column
            sheet.cell(row=row_index, column=1, value=serial_number).alignment = (
                Alignment(horizontal="center")
            )

            # Write item ID, Item Name, Description
            sheet.cell(row=row_index, column=2, value=item_id)
            sheet.cell(row=row_index, column=3, value=item_name)
            sheet.cell(row=row_index, column=4, value=description or "N/A")

            # Map quantities for each office
            col_index = 5
            total_quantity = 0
            for department, offices in department_offices.items():
                for office in offices:
                    quantity = next(
                        (
                            row["total_quantity"]
                            for row in inventory_data
                            if row["item_id__item_id"] == item_id
                            and row["office__name"] == office
                        ),
                        None,  # Use None for missing data
                    )
                    sheet.cell(row=row_index, column=col_index, value=quantity)
                    sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                        horizontal="center"
                    )
                    total_quantity += quantity or 0
                    col_index += 1

            # Write Total column
            total_cell = sheet.cell(
                row=row_index, column=col_index, value=total_quantity
            )
            total_cell.alignment = Alignment(horizontal="center")

            # Write UNIT column (empty/null)
            unit_cell = sheet.cell(row=row_index, column=col_index + 1, value=None)
            unit_cell.alignment = Alignment(horizontal="center")

            # Write Total Value column (empty if UNIT is empty)
            total_value_cell = sheet.cell(
                row=row_index, column=col_index + 2, value=None
            )
            total_value_cell.alignment = Alignment(horizontal="center")

            row_index += 1
            serial_number += 1

        # Adjust column widths
        for col in range(1, num_columns + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=broadsheet_{year}.xlsx"
        workbook.save(response)

        # Return the response
        return response
